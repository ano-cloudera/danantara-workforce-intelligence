from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import Settings


POLICY_TYPES = {"PKB", "Group Policy", "Salary Policy"}


def _split_text(text: str, size: int = 1800, overlap: int = 250) -> list[str]:
    cleaned = re.sub(r"[ \t]+", " ", text).strip()
    chunks = []
    position = 0
    while position < len(cleaned):
        chunks.append(cleaned[position : position + size])
        if position + size >= len(cleaned):
            break
        position += size - overlap
    return chunks


def _docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for paragraph in root.findall(".//w:p", namespace):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace))
        if text.strip():
            paragraphs.append(text.strip())
    return "\n".join(paragraphs)


def _xlsx_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        lines.append(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            values = [str(value) for value in row if value is not None]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


def _section(text: str) -> str | None:
    match = re.search(r"(?:Article|Pasal|Section)\s+\d+[^\n]{0,100}", text, re.I)
    return match.group(0).strip() if match else None


def load_policy_chunks(settings: Settings) -> list[dict]:
    metadata_path = settings.project_root / "data" / "workforce-app" / "demo" / "documents.json"
    if not metadata_path.exists():
        return []
    documents = json.loads(metadata_path.read_text(encoding="utf-8"))
    chunks = []
    for document in documents:
        if document.get("document_type") not in POLICY_TYPES:
            continue
        path = (settings.project_root / document["relative_path"]).resolve()
        if not path.is_file():
            continue
        pages: list[tuple[int | None, str]] = []
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            reader = PdfReader(str(path))
            pages = [
                (page_number, (page.extract_text() or "").strip())
                for page_number, page in enumerate(reader.pages, start=1)
            ]
        elif suffix == ".docx":
            pages = [(None, _docx_text(path))]
        elif suffix == ".xlsx":
            pages = [(None, _xlsx_text(path))]
        for page, text in pages:
            for index, excerpt in enumerate(_split_text(text), start=1):
                chunk_id = f'{document["document_id"]}:{page or 0}:{index}'
                chunks.append(
                    {
                        "chunk_id": chunk_id,
                        "document_id": document["document_id"],
                        "entity": document.get("entity"),
                        "title": document["title"],
                        "document_type": document["document_type"],
                        "page": page,
                        "section": _section(excerpt),
                        "text": excerpt,
                        "source_path": str(path),
                    }
                )
    return chunks
