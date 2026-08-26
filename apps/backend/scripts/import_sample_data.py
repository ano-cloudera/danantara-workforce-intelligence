#!/usr/bin/env python3
"""Build safe local demo fixtures from the supplied PoC sample package.

Raw source files stay under ``sample/``. The generated JSON contains only fields required by the
browser experience and deliberately excludes direct identifiers and protected HR attributes.
Production deployments continue to use the curated Iceberg/CDW tables through ``DATA_MODE=impala``.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - actionable CLI error
    raise SystemExit("openpyxl is required: install apps/backend/requirements.txt") from exc


ROOT = Path(__file__).resolve().parents[3]
SAMPLE = ROOT / "sample"
ADDITIONAL = SAMPLE / "additional"
RAW = SAMPLE / "data"
OUTPUT = ROOT / "data" / "workforce-app" / "demo"

GRADE_MIN_YEARS = {"G7": 12, "G5": 8, "G3": 4, "G1": 0}


def csv_rows(name: str) -> list[dict[str, str]]:
    with (ADDITIONAL / name).open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def split_csv(value: object) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def safe_document_id(path: Path) -> str:
    return re.sub(r"[^A-Z0-9]+", "-", path.stem.upper()).strip("-")


def build_candidates() -> list[dict]:
    skills: dict[str, list[dict]] = {}
    for row in csv_rows("candidate_skills.csv"):
        skills.setdefault(row["candidate_id"], []).append(
            {"name": row["skill"], "proficiency": int(row["proficiency_demo_score"])}
        )

    experiences: dict[str, list[dict]] = {}
    for row in csv_rows("candidate_experience.csv"):
        experiences.setdefault(row["candidate_id"], []).append(
            {
                "employer": row["employer"],
                "role": row["role"],
                "start_year": int(row["start_year"]),
                "end_year": int(row["end_year"]) if row["end_year"] else None,
                "summary": row["summary"],
            }
        )

    applications = {row["candidate_id"]: row for row in csv_rows("recruitment_status.csv")}
    candidates = []
    for row in csv_rows("candidate_master.csv"):
        candidate_skills = skills.get(row["candidate_id"], [])
        history = experiences.get(row["candidate_id"], [])
        application = applications.get(row["candidate_id"], {})
        summary = history[0]["summary"] if history else f'{row["current_title"]} candidate profile.'
        candidates.append(
            {
                "candidate_id": row["candidate_id"],
                "name": row["full_name"],
                "company": row["entity"],
                "current_title": row["current_title"],
                "years_experience": float(row["years_experience"]),
                "skills": [item["name"] for item in candidate_skills],
                "skill_proficiency": {
                    item["name"]: item["proficiency"] for item in candidate_skills
                },
                "summary": summary,
                "education_level": row["education_level"],
                "education_institution": row["education_institution"],
                "city": row["city"],
                "experiences": history,
                "application_id": application.get("application_id"),
                "position_id": application.get("position_id"),
                "application_stage": application.get("stage"),
                "application_status": application.get("status"),
                "salary_compliance": application.get("salary_compliance_demo"),
                "source_documents": [row["source_cv"], row["source_registration"]],
            }
        )
    return candidates


def build_positions() -> list[dict]:
    workbook = load_workbook(RAW / "job-opening.xlsx", read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    headings = [str(value) for value in rows[0]]
    positions = []
    for values in rows[1:]:
        row = dict(zip(headings, values))
        grade = str(row["grade"])
        positions.append(
            {
                "position_id": str(row["req_id"]),
                "title": str(row["position"]),
                "entity": str(row["entity"]),
                "grade": grade,
                "level": str(row["level"]),
                "department": str(row["department"]),
                "required_skills": split_csv(row["required_competencies"]),
                "preferred_skills": [],
                "min_years_experience": GRADE_MIN_YEARS.get(grade, 0),
                "competency_ids": split_csv(row["required_competency_ids"]),
                "openings": int(row["openings"] or 0),
                "status": str(row["status"]),
                "open_date": str(row["open_date"]),
            }
        )
    return positions


def build_salary_policy() -> list[dict]:
    workbook = load_workbook(RAW / "salary-policy.xlsx", read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    headings = [str(value) for value in rows[0]]
    return [dict(zip(headings, values)) for values in rows[1:]]


def document_metadata(path: Path) -> dict:
    name = path.name
    upper = name.upper()
    entity = next((code for code in ("BNS", "ENP", "NSH", "TNR") if code in upper), None)
    if "PKB" in upper:
        document_type = "PKB"
    elif "SALARY-POLICY" in upper:
        document_type = "Salary Policy"
    elif "POL" in upper or "POLICY" in upper:
        document_type = "Group Policy"
    elif "CV" in upper:
        document_type = "Candidate CV"
    elif "REGISTRATION" in upper:
        document_type = "Candidate Registration"
    elif "JOB-OPENING" in upper:
        document_type = "Job Opening"
    else:
        document_type = "Source Document"
    return {
        "document_id": safe_document_id(path),
        "title": path.stem.replace("_", " ").replace("-", " "),
        "entity": entity,
        "document_type": document_type,
        "file_name": name,
        "relative_path": str(path.relative_to(ROOT)),
        "size_bytes": path.stat().st_size,
        "downloadable": True,
    }


def write_json(name: str, value: object) -> None:
    path = OUTPUT / name
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)}")


def main() -> int:
    required = [ADDITIONAL / "candidate_master.csv", RAW / "job-opening.xlsx"]
    missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
    if missing:
        print(f"Missing sample inputs: {', '.join(missing)}", file=sys.stderr)
        return 1

    OUTPUT.mkdir(parents=True, exist_ok=True)
    write_json("candidates.json", build_candidates())
    write_json("positions.json", build_positions())
    write_json("recruitment_status.json", csv_rows("recruitment_status.csv"))
    write_json("policy_rules.json", csv_rows("policy_rules.csv"))
    write_json("salary_policy.json", build_salary_policy())
    write_json("documents.json", [document_metadata(path) for path in sorted(RAW.iterdir())])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
